"""Add a 'FAST TRACK' sheet to hand_labels_worksheet.xlsx with the 22 remaining
conversations in a stripped-down format optimized for speed.

What's compressed for speed:
  - User turns shown verbatim (the rubric scores user behavior, so user text matters most)
  - Assistant turns collapsed to first ~120 chars (enough to spot errors for NEC)
  - Sorted by total length: shortest first (quick wins build momentum)
  - 1-line decision rule per marker at the top
  - Same score columns + dropdown validation as the main sheet

Decision rules (designed to match the 0/2 pattern you've been using):
  AC  = 2 if user ASKS FOR FINISHED ANSWERS, 0 if user asks for explanation/help
  NE  = 2 if user accepts and requests next deliverable, 0 if user builds on/paraphrases
  NEC = NA (default for UltraChat — AI rarely errs noticeably). Only score 0/2 if a clear AI error appears.
  NQ  = 2 if user never asks WHY/HOW/WHAT IF, 0 if user probes reasoning
  VR  = 2 if user copies AI text or says "I'll use that"; NA if user produces no comparable own-text; 0 if paraphrases
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "validation/hand_labels_worksheet.xlsx"
POOL = ROOT / "results/cross_judge_ultrachat/conversation_pool.jsonl"


def load_pool() -> dict[str, dict]:
    pool = {}
    for line in POOL.read_text().splitlines():
        line = line.strip()
        if not line:
            continue
        r = json.loads(line)
        pool[str(r["id"])] = r
    return pool


def compact_format(turns: list[dict], assistant_max_chars: int = 120) -> str:
    """USER turns full. ASSISTANT turns truncated to first ~120 chars."""
    parts = []
    for i, t in enumerate(turns, 1):
        role = (t.get("role") or "").upper()
        content = (t.get("content") or "").strip().replace("\r", "")
        if role == "ASSISTANT" or role == "MODEL" or role == "GPT":
            short = content[:assistant_max_chars]
            if len(content) > assistant_max_chars:
                short += "  […truncated]"
            parts.append(f"  → AI ({len(content)} chars): {short}")
        else:
            parts.append(f"USER: {content}")
    return "\n".join(parts)


def main() -> None:
    wb = load_workbook(XLSX)
    ws_main = wb["Hand-coding"]
    pool = load_pool()

    # Find remaining IDs (rows where score columns D-H are all empty)
    remaining = []
    for row in range(2, 32):
        cid = ws_main.cell(row=row, column=1).value
        scores = [ws_main.cell(row=row, column=c).value for c in (4, 5, 6, 7, 8)]
        if not any(s is not None and s != "" for s in scores):
            remaining.append(str(cid))
    print(f"Remaining: {len(remaining)}: {remaining}")

    # Build records, sort by total conversation length (shortest first)
    records = []
    for cid in remaining:
        convo = pool.get(cid)
        if not convo:
            continue
        turns = convo.get("turns", [])
        total_len = sum(len(t.get("content", "")) for t in turns)
        records.append((cid, turns, total_len))
    records.sort(key=lambda x: x[2])

    # Remove existing fast-track sheet if any
    if "FAST_TRACK" in wb.sheetnames:
        del wb["FAST_TRACK"]
    ws = wb.create_sheet(title="FAST_TRACK", index=0)  # first tab

    # Styles
    bold = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    score_fill = PatternFill("solid", fgColor="FFF2CC")
    rules_fill = PatternFill("solid", fgColor="FFE7CE")
    thin = Side(border_style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_top = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center")

    # Decision-rule banner (top 8 rows)
    rules = [
        "FAST TRACK — 22 conversations left, sorted shortest-first. Aim ~30 sec/conv.",
        "",
        "Decision rules (matches your 0/2 pattern):",
        "  AC  = 2 if user asks for finished answers (\"write me…\", \"give me X\"); 0 if asks for explanation/help.",
        "  NE  = 2 if user accepts and requests next deliverable; 0 if user builds on / paraphrases.",
        "  NEC = NA (default for UltraChat). Only mark 0/2 if you see a clear AI error.",
        "  NQ  = 2 if user never asks WHY / HOW / WHAT IF; 0 if user probes reasoning.",
        "  VR  = 2 if user copies AI text or says \"I'll use that\"; NA if no comparable own-text appears; 0 if paraphrases.",
    ]
    for i, line in enumerate(rules, 1):
        c = ws.cell(row=i, column=1, value=line)
        c.fill = rules_fill
        c.alignment = wrap_top
        if i == 1:
            c.font = Font(bold=True, size=12)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=10)

    # Header row
    header_row = 10
    headers = ["id", "n_turns", "len_chars", "conversation (USER full / AI truncated)",
               "AC", "NE", "NEC", "NQ", "VR", "notes"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = bold
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    # Validations
    dv_strict = DataValidation(type="list", formula1='"0,1,2"', allow_blank=True)
    dv_lenient = DataValidation(type="list", formula1='"0,1,2,NA"', allow_blank=True)
    ws.add_data_validation(dv_strict)
    ws.add_data_validation(dv_lenient)

    # Data rows
    for i, (cid, turns, total_len) in enumerate(records):
        row = header_row + 1 + i
        ws.cell(row=row, column=1, value=cid).alignment = center
        ws.cell(row=row, column=2, value=len(turns)).alignment = center
        ws.cell(row=row, column=3, value=total_len).alignment = center
        ws.cell(row=row, column=4, value=compact_format(turns)).alignment = wrap_top

        for col in (5, 6, 7, 8, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = score_fill
            cell.alignment = center
        # NEC default: pre-fill "NA" since you've been doing that universally on UltraChat.
        # If you spot an actual AI error, just type 0 or 2 over it.
        ws.cell(row=row, column=7, value="NA")

        col_letters = {col: get_column_letter(col) for col in (5, 6, 7, 8, 9)}
        dv_strict.add(f"{col_letters[5]}{row}")
        dv_strict.add(f"{col_letters[6]}{row}")
        dv_lenient.add(f"{col_letters[7]}{row}")
        dv_strict.add(f"{col_letters[8]}{row}")
        dv_lenient.add(f"{col_letters[9]}{row}")

        for col in range(1, 11):
            ws.cell(row=row, column=col).border = border

    widths = {1: 10, 2: 8, 3: 10, 4: 110, 5: 6, 6: 6, 7: 8, 8: 6, 9: 6, 10: 22}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    for row in range(header_row + 1, header_row + 1 + len(records)):
        ws.row_dimensions[row].height = 240

    ws.freeze_panes = f"E{header_row+1}"

    wb.save(XLSX)
    print(f"Updated {XLSX}; FAST_TRACK is now the first tab.")
    print(f"NEC pre-filled as 'NA' on all 22 (override if you spot an AI error).")


if __name__ == "__main__":
    main()
