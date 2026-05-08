"""Build a polished .xlsx labeling worksheet for the 30 sampled UltraChat conversations.

Columns:
  id                    — matches conversation_pool.jsonl
  n_turns               — turn count for quick context
  conversation          — formatted multi-turn dialogue (read-only, wrapped)
  answer_copying        — fill 0/1/2
  no_elaboration        — fill 0/1/2
  no_error_correction   — fill 0/1/2/NA
  no_questioning        — fill 0/1/2
  verbatim_reuse        — fill 0/1/2/NA
  aggregate             — auto-computed from your scores (Excel formula)
  notes                 — free text for ambiguous cases

Each filled-in score column has data validation: only 0, 1, 2, or NA accepted
(NA only valid for no_error_correction and verbatim_reuse per rubric).
"""
from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
POOL = ROOT / "results/cross_judge_ultrachat/conversation_pool.jsonl"
LABEL_CSV = ROOT / "validation/human_labels_to_fill.csv"
OUT_XLSX = ROOT / "validation/hand_labels_worksheet.xlsx"


def format_conversation(turns: list[dict], max_chars_per_turn: int = 1200) -> str:
    parts = []
    for i, t in enumerate(turns, 1):
        role = (t.get("role") or "").upper()
        content = (t.get("content") or "").strip()
        if len(content) > max_chars_per_turn:
            content = content[:max_chars_per_turn] + "  […truncated…]"
        parts.append(f"[Turn {i} — {role}]\n{content}")
    return "\n\n".join(parts)


def main() -> None:
    # Load conversation pool indexed by id
    pool = {}
    for line in POOL.read_text().splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            r = json.loads(line)
            pool[str(r["id"])] = r
        except (json.JSONDecodeError, KeyError):
            continue
    print(f"Loaded {len(pool)} conversations from pool.")

    # Load shuffled order of IDs from human_labels_to_fill.csv
    ids = []
    with LABEL_CSV.open() as f:
        reader = csv.DictReader(f)
        for row in reader:
            ids.append(row["id"])
    print(f"Got {len(ids)} sampled IDs in shuffled order.")

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Hand-coding"

    headers = ["id", "n_turns", "conversation",
               "answer_copying", "no_elaboration", "no_error_correction",
               "no_questioning", "verbatim_reuse",
               "aggregate (auto)", "notes"]
    bold = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    score_fill = PatternFill("solid", fgColor="FFF2CC")  # light yellow for cells you fill
    thin = Side(border_style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = bold
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    # Data validations: 0/1/2 only for AC, NE, NQ; 0/1/2/NA for NEC and VR.
    dv_strict = DataValidation(type="list", formula1='"0,1,2"', allow_blank=True,
                                showErrorMessage=True,
                                errorTitle="Invalid score",
                                error="Use 0 (absent), 1 (mild), or 2 (strong) per rubric.")
    dv_lenient = DataValidation(type="list", formula1='"0,1,2,NA"', allow_blank=True,
                                 showErrorMessage=True,
                                 errorTitle="Invalid score",
                                 error="Use 0, 1, 2, or NA per rubric (NA = marker doesn't apply).")
    ws.add_data_validation(dv_strict)
    ws.add_data_validation(dv_lenient)

    # Fill rows
    for row_idx, cid in enumerate(ids, 2):
        convo = pool.get(cid)
        if convo is None:
            continue
        turns = convo.get("turns", [])
        n_turns = len(turns)
        text = format_conversation(turns)

        ws.cell(row=row_idx, column=1, value=cid).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=row_idx, column=2, value=n_turns).alignment = Alignment(horizontal="center", vertical="top")
        c = ws.cell(row=row_idx, column=3, value=text)
        c.alignment = Alignment(wrap_text=True, vertical="top")

        # Score cells (yellow)
        for col in (4, 5, 6, 7, 8):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = score_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Apply validations
        col_letters = {col: get_column_letter(col) for col in (4, 5, 6, 7, 8)}
        dv_strict.add(f"{col_letters[4]}{row_idx}")  # answer_copying
        dv_strict.add(f"{col_letters[5]}{row_idx}")  # no_elaboration
        dv_lenient.add(f"{col_letters[6]}{row_idx}")  # no_error_correction (NA allowed)
        dv_strict.add(f"{col_letters[7]}{row_idx}")  # no_questioning
        dv_lenient.add(f"{col_letters[8]}{row_idx}")  # verbatim_reuse (NA allowed)

        # Aggregate formula: AVERAGEIF over numeric scores (skips NA)
        # IFERROR wraps in case all are NA
        formula = (f'=IFERROR(AVERAGE(IF(ISNUMBER(D{row_idx}:H{row_idx}),'
                   f'D{row_idx}:H{row_idx})), "")')
        ws.cell(row=row_idx, column=9, value=formula).alignment = Alignment(horizontal="center")

        # Notes column: empty for free text
        ws.cell(row=row_idx, column=10, value="").alignment = Alignment(wrap_text=True, vertical="top")

        # Borders for all columns this row
        for col in range(1, 11):
            ws.cell(row=row_idx, column=col).border = border

    # Column widths
    widths = {1: 22, 2: 8, 3: 90,
              4: 14, 5: 14, 6: 18, 7: 14, 8: 14,
              9: 14, 10: 30}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Row heights: header taller, data rows tall enough to show conversation
    ws.row_dimensions[1].height = 36
    for row_idx in range(2, len(ids) + 2):
        ws.row_dimensions[row_idx].height = 360  # tall row; user can resize

    # Freeze header + id columns
    ws.freeze_panes = "D2"

    # Add a "rubric quick reference" sheet
    ws2 = wb.create_sheet(title="Rubric reference")
    rubric = [
        ["Marker", "0 (absent)", "1 (mild)", "2 (strong)", "NA allowed?"],
        ["answer_copying",
         "User asks for explanation/help/verification.",
         "Asks for answer with some context (\"got X, is that right?\").",
         "User asks for finished outputs (\"write me a 500-word essay\").",
         "no"],
        ["no_elaboration",
         "User builds on, paraphrases, applies AI output.",
         "Brief acknowledgement (\"thanks, ok\") then topic change.",
         "Accepts and immediately requests next deliverable.",
         "no"],
        ["no_error_correction",
         "User catches errors / asks clarification.",
         "Notices but doesn't investigate.",
         "AI made an error and user accepted uncritically.",
         "yes — when AI response has no apparent issues to correct"],
        ["no_questioning",
         "User asks why / how / probes reasoning.",
         "User asks only logistical (reformat, shorter, in bullets).",
         "User asks only for more outputs, never for understanding.",
         "no"],
        ["verbatim_reuse",
         "User text is independent / paraphrases.",
         "Borrows phrases but mostly own.",
         "Largely identical to AI output, OR explicit reuse statement.",
         "yes — when user produces no comparable own-text"],
    ]
    for row_idx, row in enumerate(rubric, 1):
        for col_idx, val in enumerate(row, 1):
            c = ws2.cell(row=row_idx, column=col_idx, value=val)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if row_idx == 1:
                c.font = bold
                c.fill = header_fill
            c.border = border
    for col, w in {1: 22, 2: 38, 3: 38, 4: 38, 5: 28}.items():
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[1].height = 30
    for row_idx in range(2, 7):
        ws2.row_dimensions[row_idx].height = 90

    # Add a quick-instructions sheet
    ws3 = wb.create_sheet(title="Instructions", index=0)  # make it the first tab
    instr = [
        ["Hand-coding instructions"],
        [],
        ["1. Go to the 'Hand-coding' tab and read each conversation in column C."],
        ["2. For each row, fill the 5 yellow score columns (D-H):"],
        ["     - answer_copying, no_elaboration, no_questioning: 0, 1, or 2"],
        ["     - no_error_correction, verbatim_reuse: 0, 1, 2, or NA"],
        ["3. The 'Rubric reference' tab has the scoring anchors. The full rubric is in rubric.md at the project root."],
        ["4. The 'aggregate (auto)' column auto-computes mean over numeric scores (NA is excluded)."],
        ["5. Use the 'notes' column for ambiguous cases. These help the writeup later."],
        [],
        ["Tips:"],
        ["  - Score the USER's behavior, not the assistant's."],
        ["  - When unsure between two scores, default to the LOWER integer."],
        ["  - Try to label all 30 in one sitting; drift across days is real."],
        ["  - Don't look at the LLM-grader scores until you're done (they're in the JSONL files in results/cross_judge_ultrachat/, but ignore them while labeling)."],
        [],
        ["When done:"],
        ["  - Save (Cmd+S) — the .xlsx already retains your scores."],
        ["  - To compute kappa vs each LLM judge, export this sheet to validation/human_labels.csv with the columns:"],
        ["       id, answer_copying, no_elaboration, no_error_correction, no_questioning, verbatim_reuse, notes"],
        ["    Then run the validate.py commands in the README."],
    ]
    for row_idx, row in enumerate(instr, 1):
        for col_idx, val in enumerate(row, 1):
            c = ws3.cell(row=row_idx, column=col_idx, value=val)
            if row_idx == 1:
                c.font = Font(bold=True, size=14)
            c.alignment = Alignment(vertical="top", wrap_text=True)
    ws3.column_dimensions["A"].width = 110

    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
